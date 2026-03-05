import io
import datetime as dt
import re
import unicodedata

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


st.set_page_config(page_title="Fechamento Afiliados", layout="wide")
st.title("Fechamento mensal de comissões de afiliados")
st.caption(
    "Fluxo: subir Plan Base (e-commerce) + Plan Afiliados → baixar retorno no layout do Modelo "
    "(sem precisar enviar o Modelo.xlsx)."
)

# ===================== CONFIG (SEUS CABEÇALHOS) =====================
# Plan Base (e-commerce)
ECOM_ORDER_COL = "Order2"
ECOM_STATUS_COL = "Status"
ECOM_FREIGHT_COL = "Shipping Value"     # frete RATEADO -> SUM por pedido
ECOM_ORDER_VALUE_COL = "Total Value"    # total do pedido -> MAX por pedido
ECOM_EMISSION_DATE_COL = "Creation D"   # data de emissão na base

# Plan Afiliados (chave)
AFF_ORDER_COL = "Order ID"

# Nome da aba do arquivo final (igual ao seu Modelo.xlsx)
OUTPUT_SHEET_NAME = "Fevereiro (2)"

# Colunas finais EXACTAS do Modelo.xlsx (ordem)
MODEL_HEADERS = [
    "Order ID",
    "Valor líquido",
    "Comissão",
    "Data",
    "Afiliado",
    "Device",
    "Status",
    "Frete",
    "Valor Vtex",
    "Valor S/ frete",
    "Motivo de Cancelamento",
    "Captada Vitrio",
]

# Larguras baseadas no seu Modelo.xlsx
COL_WIDTHS = {
    "A": 16.140625,
    "B": 13.5703125,
    "C": 11.85546875,
    "D": 17.5703125,
    "E": 47.42578125,
    "F": 7.0,
    "G": 22.7109375,
    "H": 13.0,
    "I": 13.0,
    "J": 13.0,
    "K": 23.85546875,
    "L": 13.0,
}

# Formatos
FMT_ORDER = "00000"
FMT_CURRENCY_BR = '_-"R$"\\ * #,##0.00_-;\\-"R$"\\ * #,##0.00_-;_-"R$"\\ * "-"??_-;_-@'
FMT_DATETIME = "m/d/yy h:mm"
# ====================================================================


# --------------------- Helpers ---------------------
def normalize_text(x) -> str:
    s = str(x) if x is not None else ""
    s = s.strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s)
    return s

def read_any(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    return pd.read_excel(uploaded_file)

def drop_blank_technical_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Remove Unnamed:*, nomes numéricos (0,1,2...) e colunas 100% vazias."""
    cols_to_drop = []
    for c in df.columns:
        cs = str(c)
        if cs.startswith("Unnamed:"):
            cols_to_drop.append(c)
        elif re.fullmatch(r"\d+", cs):
            cols_to_drop.append(c)
    df = df.drop(columns=cols_to_drop, errors="ignore")

    empty_cols = [c for c in df.columns if df[c].isna().all()]
    df = df.drop(columns=empty_cols, errors="ignore")
    return df

def ensure_cols(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    for c in cols:
        if c not in df.columns:
            df[c] = pd.NA
    return df

def build_output_workbook(df_out: pd.DataFrame) -> bytes:
    """Gera um xlsx com a mesma estrutura do Modelo (colunas/ordem/aba/larguras/formatos)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = OUTPUT_SHEET_NAME

    # Estilos (básicos, consistentes)
    header_font = Font(b=True, size=11, name="Aptos Narrow")
    header_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="9E9E9E")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    for col_idx, col_name in enumerate(MODEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = header_fill
        cell.border = header_border

    # Larguras
    for col_letter, w in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = w

    # Dados (somente colunas do modelo e na ordem)
    df_out = ensure_cols(df_out, MODEL_HEADERS)[MODEL_HEADERS].copy()

    for r, (_, row) in enumerate(df_out.iterrows(), start=2):
        for c, col_name in enumerate(MODEL_HEADERS, start=1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            ws.cell(r, c, val)

    # Formatos
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        ws.cell(r, 1).number_format = FMT_ORDER      # Order ID
        ws.cell(r, 2).number_format = FMT_CURRENCY_BR  # Valor líquido
        ws.cell(r, 3).number_format = FMT_CURRENCY_BR  # Comissão
        ws.cell(r, 4).number_format = FMT_DATETIME     # Data
        ws.cell(r, 8).number_format = FMT_CURRENCY_BR  # Frete
        ws.cell(r, 9).number_format = FMT_CURRENCY_BR  # Valor Vtex
        ws.cell(r, 10).number_format = FMT_CURRENCY_BR # Valor S/ frete

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------- Parâmetros do fechamento ---------------------
st.subheader("Parâmetros do fechamento (regra do corte)")

ref_month = st.date_input(
    "Mês de referência do fechamento (qualquer dia dentro do mês)",
    value=dt.date.today().replace(day=1)
)
cutoff_day = st.number_input("Dia limite (corte)", min_value=1, max_value=31, value=20, step=1)
cutoff_date = dt.date(ref_month.year, ref_month.month, int(cutoff_day))
cutoff_ts = pd.Timestamp(cutoff_date)

st.caption(f"Corte: pedidos com emissão até {cutoff_date} e status selecionado serão marcados como Cancelado.")

# --------------------- Upload ---------------------
col1, col2 = st.columns(2)
with col1:
    ecom_file = st.file_uploader("📄 Plan Base (e-commerce) - .csv ou .xlsx", type=["csv", "xlsx"])
with col2:
    aff_file = st.file_uploader("📄 Plan Afiliados - .csv ou .xlsx", type=["csv", "xlsx"])

if not (ecom_file and aff_file):
    st.warning("Suba os 2 arquivos para processar e gerar o fechamento.")
    st.stop()

# --------------------- Leitura ---------------------
try:
    ecom = read_any(ecom_file)
    aff = read_any(aff_file)
except Exception as e:
    st.error(f"Erro ao ler os arquivos: {e}")
    st.stop()

# Remove colunas lixo/vazias do Afiliados para evitar “Unnamed” no resultado
aff = drop_blank_technical_columns(aff)

# --------------------- Validação Base ---------------------
needed_ecom = {ECOM_ORDER_COL, ECOM_STATUS_COL, ECOM_FREIGHT_COL, ECOM_ORDER_VALUE_COL, ECOM_EMISSION_DATE_COL}
missing_ecom = needed_ecom - set(ecom.columns)
if missing_ecom:
    st.error(
        f"Faltam colunas na Plan Base: {sorted(list(missing_ecom))}\n\n"
        f"Colunas existentes: {list(ecom.columns)}"
    )
    st.stop()

# Garante que a planilha de afiliados tenha as colunas do modelo (se não tiver, cria vazio)
aff = ensure_cols(aff, MODEL_HEADERS)

# --------------------- Normalização / parsing ---------------------
ecom[ECOM_ORDER_COL] = ecom[ECOM_ORDER_COL].astype(str).str.strip()
aff[AFF_ORDER_COL] = aff[AFF_ORDER_COL].astype(str).str.strip()

ecom[ECOM_FREIGHT_COL] = pd.to_numeric(ecom[ECOM_FREIGHT_COL], errors="coerce").fillna(0)
ecom[ECOM_ORDER_VALUE_COL] = pd.to_numeric(ecom[ECOM_ORDER_VALUE_COL], errors="coerce").fillna(0)

ecom[ECOM_EMISSION_DATE_COL] = pd.to_datetime(
    ecom[ECOM_EMISSION_DATE_COL],
    errors="coerce",
    dayfirst=True
)

# Status alvo via dropdown (valores reais da base)
unique_status = sorted([s for s in ecom[ECOM_STATUS_COL].dropna().astype(str).unique()])
default_idx = unique_status.index("Preparando Entrega") if "Preparando Entrega" in unique_status else 0
status_target_value = st.selectbox(
    "Qual status deve virar CANCELADO quando emissão <= corte?",
    options=unique_status,
    index=default_idx
)
status_target_norm = normalize_text(status_target_value)

status_forced_value = st.text_input("Status aplicado pela regra (valor reportado)", value="Cancelado").strip()

# --------------------- Consolidação Base por pedido ---------------------
# Frete rateado -> SUM
freight_by_order = (
    ecom.groupby(ECOM_ORDER_COL, as_index=False)[ECOM_FREIGHT_COL]
    .sum()
    .rename(columns={ECOM_FREIGHT_COL: "frete_consolidado"})
)

# Total do pedido -> MAX (evita duplicação)
value_by_order = (
    ecom.groupby(ECOM_ORDER_COL, as_index=False)[ECOM_ORDER_VALUE_COL]
    .max()
    .rename(columns={ECOM_ORDER_VALUE_COL: "valor_pedido_consolidado"})
)

# Emissão -> MIN
emissao_by_order = (
    ecom.groupby(ECOM_ORDER_COL, as_index=False)[ECOM_EMISSION_DATE_COL]
    .min()
    .rename(columns={ECOM_EMISSION_DATE_COL: "data_emissao_consolidada"})
)

# Status -> primeiro não nulo
status_by_order = (
    ecom[[ECOM_ORDER_COL, ECOM_STATUS_COL]]
    .dropna(subset=[ECOM_STATUS_COL])
    .groupby(ECOM_ORDER_COL, as_index=False)[ECOM_STATUS_COL]
    .first()
    .rename(columns={ECOM_STATUS_COL: "status_pedido"})
)

cons = status_by_order.merge(freight_by_order, how="outer", on=ECOM_ORDER_COL)
cons = cons.merge(value_by_order, how="outer", on=ECOM_ORDER_COL)
cons = cons.merge(emissao_by_order, how="outer", on=ECOM_ORDER_COL)

cons["frete_consolidado"] = cons["frete_consolidado"].fillna(0)
cons["valor_pedido_consolidado"] = cons["valor_pedido_consolidado"].fillna(0)
cons["status_pedido"] = cons["status_pedido"].fillna("Pedido não encontrado na base")

# --------------------- Regra do corte (status -> cancelado) ---------------------
cons["status_norm"] = cons["status_pedido"].map(normalize_text)
mask_regra = (
    (cons["status_norm"] == status_target_norm) &
    (cons["data_emissao_consolidada"].notna()) &
    (cons["data_emissao_consolidada"] <= cutoff_ts)
)
cons.loc[mask_regra, "status_pedido"] = status_forced_value
cons.drop(columns=["status_norm"], inplace=True, errors="ignore")

# --------------------- Enriquecimento do Afiliados ---------------------
tmp = cons.rename(columns={ECOM_ORDER_COL: AFF_ORDER_COL}).copy()

# >>> AJUSTE PRINCIPAL PEDIDO PELO CELSO:
# Valor S/ frete deve ser calculado para TODOS os status (não zera por status).
tmp["valor_s_frete"] = (tmp["valor_pedido_consolidado"] - tmp["frete_consolidado"]).clip(lower=0)

enriq = aff.merge(
    tmp[[AFF_ORDER_COL, "status_pedido", "frete_consolidado", "valor_pedido_consolidado", "valor_s_frete"]],
    on=AFF_ORDER_COL,
    how="left"
)

# Preenche colunas do “Modelo”
enriq["Status"] = enriq["status_pedido"]
enriq["Frete"] = enriq["frete_consolidado"]
enriq["Valor Vtex"] = enriq["valor_pedido_consolidado"]
enriq["Valor S/ frete"] = enriq["valor_s_frete"]

# Não mexe na coluna "Comissão" (fica como vem do afiliados)
# Motivo de Cancelamento: mantém o que vier; se vazio, continua vazio.

# Remove colunas auxiliares
enriq.drop(
    columns=["status_pedido", "frete_consolidado", "valor_pedido_consolidado", "valor_s_frete"],
    inplace=True,
    errors="ignore"
)

# Garante somente colunas do modelo e sem lixo
saida = ensure_cols(enriq, MODEL_HEADERS)[MODEL_HEADERS].copy()
saida = drop_blank_technical_columns(saida)

# --------------------- Preview / Métricas ---------------------
st.subheader("Prévia do retorno (primeiras 50 linhas) — layout Modelo")
st.dataframe(saida.head(50), use_container_width=True)

st.info(
    f"Pedidos forçados para '{status_forced_value}' pela regra: {int(mask_regra.sum())} | "
    f"Status alvo selecionado: {status_target_value} | Corte: {cutoff_date}"
)

# --------------------- Download ---------------------
xlsx_bytes = build_output_workbook(saida)

st.download_button(
    "⬇️ Baixar planilha final (layout do Modelo)",
    data=xlsx_bytes,
    file_name="fechamento_afiliados_modelo.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
